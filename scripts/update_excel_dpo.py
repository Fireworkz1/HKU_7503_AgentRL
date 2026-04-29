import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
import json
import copy

EXCEL_PATH = '../eval_results/training_summary.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH)

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# 1. Add DPO训练日志 sheet (DPO v1 - 1 epoch, lr=4e-8)
# ============================================================
def parse_dpo_log(log_path):
    rows = []
    with open(log_path, 'r', encoding='utf-8') as f:
        for line in f:
            m = re.search(
                r'Epoch:\[(\d+)/(\d+)\]\((\d+)/(\d+)\),\s*loss:\s*([\d.]+),\s*dpo_loss:\s*([\d.]+),\s*aux_loss:\s*([\d.]+),\s*learning_rate:\s*([\d.]+)',
                line
            )
            if m:
                epoch = int(m.group(1))
                total_epochs = int(m.group(2))
                step = int(m.group(3))
                total_steps = int(m.group(4))
                loss = float(m.group(5))
                dpo_loss = float(m.group(6))
                aux_loss = float(m.group(7))
                lr = float(m.group(8))
                global_step = (epoch - 1) * total_steps + step
                rows.append([global_step, epoch, step, loss, dpo_loss, aux_loss, lr])
    return rows

def create_dpo_log_sheet(wb, sheet_name, log_path):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = ['Global_Step', 'Epoch', 'Step', 'Loss', 'DPO_Loss', 'Aux_Loss', 'Learning_Rate']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    rows = parse_dpo_log(log_path)
    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center_align
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    print(f"  Created sheet '{sheet_name}' with {len(rows)} rows")

# DPO v1
create_dpo_log_sheet(wb, 'DPO训练日志(v1)', '../logs/dpo_train.log')
# DPO v2
create_dpo_log_sheet(wb, 'DPO训练日志(v2)', '../logs/dpo_train_v2.log')

# ============================================================
# 2. Update ToolCall评测(80题) sheet - add DPO columns
# ============================================================
toolcall_sheet_name = wb.sheetnames[3]  # ToolCall评测(80题)
ws_tc = wb[toolcall_sheet_name]

# Read existing headers
existing_headers = []
for cell in ws_tc[1]:
    existing_headers.append(cell.value)

# Copy existing header style
sample_header_cell = ws_tc.cell(row=1, column=1)

# DPO v1 eval data
with open('../eval_results/toolcall_batch_dpo.json', 'r') as f:
    dpo_v1_eval = json.load(f)

# DPO v2 eval data
with open('../eval_results/toolcall_batch_dpo_v2.json', 'r') as f:
    dpo_v2_eval = json.load(f)

# Category mapping (order in the sheet)
cat_order = ['math_single', 'single_other', 'multi_step', 'tool_select', 'lang_variant']

# Current last column
last_col = ws_tc.max_column

# Add DPO v1 columns (4 columns: 调用工具, 工具正确, 格式完整, 完成率)
dpo_v1_headers = ['DPO_v1 调用工具', 'DPO_v1 工具正确', 'DPO_v1 格式完整', 'DPO_v1 完成率']
for i, h in enumerate(dpo_v1_headers):
    col = last_col + 1 + i
    cell = ws_tc.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

# Add DPO v2 columns
dpo_v2_headers = ['DPO_v2 调用工具', 'DPO_v2 工具正确', 'DPO_v2 格式完整', 'DPO_v2 完成率']
for i, h in enumerate(dpo_v2_headers):
    col = last_col + 5 + i
    cell = ws_tc.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

# Fill DPO v1 data per category
for row_idx, cat in enumerate(cat_order, 2):
    stats = dpo_v1_eval['category_stats'][cat]
    total = stats['total']
    called_any = stats['called_any']
    called_correct = stats['called_correct']
    chain_complete = stats['chain_complete']
    rate = f"{chain_complete/total*100:.1f}%"

    for ci, val in enumerate([called_any, called_correct, chain_complete, rate]):
        cell = ws_tc.cell(row=row_idx, column=last_col + 1 + ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border

# DPO v1 totals row
total_row = len(cat_order) + 2  # row 7
for ci, val in enumerate([
    dpo_v1_eval['called_any'], dpo_v1_eval['called_correct'],
    dpo_v1_eval['chain_complete'],
    f"{dpo_v1_eval['chain_complete']/dpo_v1_eval['total']*100:.1f}%"
]):
    cell = ws_tc.cell(row=total_row, column=last_col + 1 + ci, value=val)
    cell.alignment = center_align
    cell.border = thin_border
    cell.font = Font(bold=True)

# Fill DPO v2 data per category
for row_idx, cat in enumerate(cat_order, 2):
    stats = dpo_v2_eval['category_stats'][cat]
    total = stats['total']
    called_any = stats['called_any']
    called_correct = stats['called_correct']
    chain_complete = stats['chain_complete']
    rate = f"{chain_complete/total*100:.1f}%"

    for ci, val in enumerate([called_any, called_correct, chain_complete, rate]):
        cell = ws_tc.cell(row=row_idx, column=last_col + 5 + ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border

# DPO v2 totals row
for ci, val in enumerate([
    dpo_v2_eval['called_any'], dpo_v2_eval['called_correct'],
    dpo_v2_eval['chain_complete'],
    f"{dpo_v2_eval['chain_complete']/dpo_v2_eval['total']*100:.1f}%"
]):
    cell = ws_tc.cell(row=total_row, column=last_col + 5 + ci, value=val)
    cell.alignment = center_align
    cell.border = thin_border
    cell.font = Font(bold=True)

# Adjust column widths
for col in range(last_col + 1, last_col + 9):
    ws_tc.column_dimensions[get_column_letter(col)].width = 16

print(f"  Updated '{toolcall_sheet_name}' with DPO v1 & v2 columns")

# ============================================================
# 3. Update 训练对比汇总 sheet - add DPO columns
# ============================================================
compare_sheet_name = wb.sheetnames[2]  # 训练对比汇总
ws_cmp = wb[compare_sheet_name]

# Current columns: 指标, PPO 初始, PPO 最终, PPO 变化, GRPO 初始, GRPO 最终, GRPO 变化
# Add: DPO_v1 初始, DPO_v1 最终, DPO_v1 变化, DPO_v2 初始, DPO_v2 最终, DPO_v2 变化
cmp_last_col = ws_cmp.max_column

# DPO v1 log data
dpo_v1_rows = parse_dpo_log('../logs/dpo_train.log')
dpo_v1_first_loss = dpo_v1_rows[0][3]   # loss
dpo_v1_last_loss = dpo_v1_rows[-1][3]
dpo_v1_first_lr = dpo_v1_rows[0][6]
dpo_v1_last_lr = dpo_v1_rows[-1][6]

# DPO v2 log data
dpo_v2_rows = parse_dpo_log('../logs/dpo_train_v2.log')
dpo_v2_first_loss = dpo_v2_rows[0][3]
dpo_v2_last_loss = dpo_v2_rows[-1][3]
dpo_v2_first_lr = dpo_v2_rows[0][6]
dpo_v2_last_lr = dpo_v2_rows[-1][6]

# Add DPO v1 headers
dpo_v1_cmp_headers = ['DPO_v1 初始', 'DPO_v1 最终', 'DPO_v1 变化']
for i, h in enumerate(dpo_v1_cmp_headers):
    col = cmp_last_col + 1 + i
    cell = ws_cmp.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

# Add DPO v2 headers
dpo_v2_cmp_headers = ['DPO_v2 初始', 'DPO_v2 最终', 'DPO_v2 变化']
for i, h in enumerate(dpo_v2_cmp_headers):
    col = cmp_last_col + 4 + i
    cell = ws_cmp.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border

# DPO metrics: DPO_Loss, Learning_Rate, ToolCall完成率
# Row 2: DPO_Loss (analogous to Reward)
# Row 3: Learning_Rate (analogous to KL_ref)
# Row 4: ToolCall完成率
# Row 5: (empty or additional)

# Clear existing rows 2-5 for DPO columns and rewrite
# Row 2: DPO_Loss
dpo_metrics = [
    ('DPO_Loss', dpo_v1_first_loss, dpo_v1_last_loss, dpo_v2_first_loss, dpo_v2_last_loss),
    ('Learning_Rate', dpo_v1_first_lr, dpo_v1_last_lr, dpo_v2_first_lr, dpo_v2_last_lr),
    ('ToolCall完成率', '—', f"{dpo_v1_eval['chain_complete']/dpo_v1_eval['total']*100:.1f}%", '—', f"{dpo_v2_eval['chain_complete']/dpo_v2_eval['total']*100:.1f}%"),
]

for r_offset, (metric_name, v1_start, v1_end, v2_start, v2_end) in enumerate(dpo_metrics):
    row = r_offset + 2
    # DPO v1
    for ci, val in enumerate([v1_start, v1_end]):
        cell = ws_cmp.cell(row=row, column=cmp_last_col + 1 + ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border
    # DPO v1 change
    if isinstance(v1_start, (int, float)) and isinstance(v1_end, (int, float)):
        change = round(v1_end - v1_start, 4)
    else:
        change = '—'
    cell = ws_cmp.cell(row=row, column=cmp_last_col + 3, value=change)
    cell.alignment = center_align
    cell.border = thin_border

    # DPO v2
    for ci, val in enumerate([v2_start, v2_end]):
        cell = ws_cmp.cell(row=row, column=cmp_last_col + 4 + ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border
    # DPO v2 change
    if isinstance(v2_start, (int, float)) and isinstance(v2_end, (int, float)):
        change = round(v2_end - v2_start, 4)
    else:
        change = '—'
    cell = ws_cmp.cell(row=row, column=cmp_last_col + 6, value=change)
    cell.alignment = center_align
    cell.border = thin_border

# Adjust column widths
for col in range(cmp_last_col + 1, cmp_last_col + 7):
    ws_cmp.column_dimensions[get_column_letter(col)].width = 15

print(f"  Updated '{compare_sheet_name}' with DPO comparison data")

# ============================================================
# 4. Update 训练曲线 sheet - add DPO loss curves
# ============================================================
curve_sheet_name = wb.sheetnames[5]  # 训练曲线
ws_curve = wb[curve_sheet_name]

curve_last_col = ws_curve.max_column

# Add DPO v1 columns
col_v1_step = curve_last_col + 2  # leave a gap column
col_v1_loss = curve_last_col + 3

cell = ws_curve.cell(row=1, column=col_v1_step, value='DPO_v1_Step')
cell.font = header_font_white
cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
cell.alignment = center_align
cell.border = thin_border

cell = ws_curve.cell(row=1, column=col_v1_loss, value='DPO_v1_Loss')
cell.font = header_font_white
cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
cell.alignment = center_align
cell.border = thin_border

for r_idx, row_data in enumerate(dpo_v1_rows, 2):
    global_step = row_data[0]
    loss = row_data[3]
    cell = ws_curve.cell(row=r_idx, column=col_v1_step, value=global_step)
    cell.alignment = center_align
    cell.border = thin_border
    cell = ws_curve.cell(row=r_idx, column=col_v1_loss, value=loss)
    cell.alignment = center_align
    cell.border = thin_border

# Add DPO v2 columns
col_v2_step = col_v1_loss + 2  # leave a gap column
col_v2_loss = col_v1_loss + 3

cell = ws_curve.cell(row=1, column=col_v2_step, value='DPO_v2_Step')
cell.font = header_font_white
cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
cell.alignment = center_align
cell.border = thin_border

cell = ws_curve.cell(row=1, column=col_v2_loss, value='DPO_v2_Loss')
cell.font = header_font_white
cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
cell.alignment = center_align
cell.border = thin_border

for r_idx, row_data in enumerate(dpo_v2_rows, 2):
    global_step = row_data[0]
    loss = row_data[3]
    cell = ws_curve.cell(row=r_idx, column=col_v2_step, value=global_step)
    cell.alignment = center_align
    cell.border = thin_border
    cell = ws_curve.cell(row=r_idx, column=col_v2_loss, value=loss)
    cell.alignment = center_align
    cell.border = thin_border

# Adjust column widths
for col in [col_v1_step, col_v1_loss, col_v2_step, col_v2_loss]:
    ws_curve.column_dimensions[get_column_letter(col)].width = 15

print(f"  Updated '{curve_sheet_name}' with DPO loss curves")

# ============================================================
# 5. Add DPO训练配置 sheet
# ============================================================
config_sheet_name = 'DPO训练配置'
if config_sheet_name in wb.sheetnames:
    del wb[config_sheet_name]
ws_cfg = wb.create_sheet(config_sheet_name)

config_headers = ['参数', 'DPO_v1', 'DPO_v2']
for ci, h in enumerate(config_headers, 1):
    cell = ws_cfg.cell(row=1, column=ci, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

config_data = [
    ('算法', 'DPO', 'DPO'),
    ('模型参数量', '63.912M', '63.912M'),
    ('基础权重', 'full_sft', 'full_sft'),
    ('训练数据', 'dpo.jsonl (17166 pairs)', 'dpo.jsonl (17166 pairs)'),
    ('Epochs', 1, 2),
    ('Batch Size', 4, 4),
    ('Learning Rate', '4e-8', '5e-7'),
    ('Beta', 0.15, 0.15),
    ('Max Seq Len', 1024, 1024),
    ('Hidden Size', 768, 768),
    ('Num Hidden Layers', 8, 8),
    ('Grad Clip', 1.0, 1.0),
    ('Dtype', 'bfloat16', 'bfloat16'),
    ('Optimizer', 'AdamW', 'AdamW'),
    ('LR Schedule', 'Cosine', 'Cosine'),
    ('Total Steps', 4292, 8584),
    ('初始Loss', dpo_v1_first_loss, dpo_v2_first_loss),
    ('最终Loss', dpo_v1_last_loss, dpo_v2_last_loss),
    ('权重文件', 'dpo_768_ours.pth', 'dpo_v2_768.pth'),
    ('ToolCall完成率', f"{dpo_v1_eval['chain_complete']/dpo_v1_eval['total']*100:.1f}%",
     f"{dpo_v2_eval['chain_complete']/dpo_v2_eval['total']*100:.1f}%"),
]

for r_idx, (param, v1, v2) in enumerate(config_data, 2):
    for ci, val in enumerate([param, v1, v2], 1):
        cell = ws_cfg.cell(row=r_idx, column=ci, value=val)
        cell.alignment = center_align
        cell.border = thin_border
        if ci == 1:
            cell.font = Font(bold=True)

ws_cfg.column_dimensions['A'].width = 20
ws_cfg.column_dimensions['B'].width = 25
ws_cfg.column_dimensions['C'].width = 25

print(f"  Created '{config_sheet_name}' sheet")

# ============================================================
# Save
# ============================================================
wb.save(EXCEL_PATH)
print(f"\nSaved to {EXCEL_PATH}")
