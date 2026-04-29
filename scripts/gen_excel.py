import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

wb = Workbook()

# ==================== Sheet1: PPO 训练日志 ====================
ws_ppo = wb.active
ws_ppo.title = "PPO训练日志"
ppo_headers = ["Step", "Reward", "KL_ref", "Approx_KL", "ClipFrac", "Critic_Loss", "Avg_Response_Len"]
for col, h in enumerate(ppo_headers, 1):
    c = ws_ppo.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    c.alignment = Alignment(horizontal="center")

ppo_data = []
with open("ppo_train.log", "r", encoding="utf-8") as f:
    for line in f:
        m = re.search(r"Epoch:\[1/1\]\((\d+)/9751\), Reward: ([-\d.]+), KL_ref: ([-\d.]+), Approx KL: ([-\d.]+), ClipFrac: ([-\d.]+), Critic Loss: ([-\d.]+), Avg Response Len: ([-\d.]+)", line)
        if m:
            row = [int(m.group(1)), float(m.group(2)), float(m.group(3)), float(m.group(4)), float(m.group(5)), float(m.group(6)), float(m.group(7))]
            ppo_data.append(row)

# 每10步采样
for row in ppo_data:
    if row[0] % 10 == 0 or row[0] <= 10 or row[0] >= 9740:
        ws_ppo.append(row)

for col in range(1, 8):
    ws_ppo.column_dimensions[get_column_letter(col)].width = 16

# ==================== Sheet2: GRPO 训练日志 ====================
ws_grpo = wb.create_sheet("GRPO训练日志")
grpo_headers = ["Step", "Reward", "KL_ref", "Adv_Std", "Adv_Mean", "Actor_Loss", "Avg_Response_Len"]
for col, h in enumerate(grpo_headers, 1):
    c = ws_grpo.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    c.alignment = Alignment(horizontal="center")

grpo_data = []
with open("grpo_sglang.log", "r", encoding="utf-8") as f:
    for line in f:
        m = re.search(r"Epoch:\[1/1\]\((\d+)/9751\), Reward: ([-\d.]+), KL_ref: ([-\d.]+), Adv Std: ([-\d.]+), Adv Mean: ([-\d.]+), Actor Loss: ([-\d.]+), Avg Response Len: ([-\d.]+)", line)
        if m:
            row = [int(m.group(1)), float(m.group(2)), float(m.group(3)), float(m.group(4)), float(m.group(5)), float(m.group(6)), float(m.group(7))]
            grpo_data.append(row)

for row in grpo_data:
    ws_grpo.append(row)

for col in range(1, 8):
    ws_grpo.column_dimensions[get_column_letter(col)].width = 16

# ==================== Sheet3: 训练对比汇总 ====================
ws_summary = wb.create_sheet("训练对比汇总")
header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

summary_headers = ["指标", "PPO 起始", "PPO 结束", "PPO 变化", "GRPO 起始", "GRPO 结束", "GRPO 变化"]
for col, h in enumerate(summary_headers, 1):
    c = ws_summary.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

ppo_first, ppo_last = ppo_data[0], ppo_data[-1]
grpo_first, grpo_last = grpo_data[0], grpo_data[-1]

summary_rows = [
    ["Reward", ppo_first[1], ppo_last[1], round(ppo_last[1]-ppo_first[1], 4), grpo_first[1], grpo_last[1], round(grpo_last[1]-grpo_first[1], 4)],
    ["KL_ref", ppo_first[2], ppo_last[2], round(ppo_last[2]-ppo_first[2], 4), grpo_first[2], grpo_last[2], round(grpo_last[2]-grpo_first[2], 4)],
    ["Avg_Response_Len", ppo_first[6], ppo_last[6], round(ppo_last[6]-ppo_first[6], 2), grpo_first[6], grpo_last[6], round(grpo_last[6]-grpo_first[6], 2)],
    ["Loss(Critic/Actor)", ppo_first[5], ppo_last[5], round(ppo_last[5]-ppo_first[5], 4), grpo_first[5], grpo_last[5], round(grpo_last[5]-grpo_first[5], 4)],
]
for row in summary_rows:
    ws_summary.append(row)

for col in range(1, 8):
    ws_summary.column_dimensions[get_column_letter(col)].width = 20

# ==================== Sheet4: ToolCall 评测结果 ====================
ws_eval = wb.create_sheet("ToolCall评测(80题)")
eval_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
eval_headers = ["类别", "题数", "SFT 调用工具", "SFT 工具正确", "SFT 链式完整", "SFT 完整率",
                "GRPO 调用工具", "GRPO 工具正确", "GRPO 链式完整", "GRPO 完整率",
                "Agent 调用工具", "Agent 工具正确", "Agent 链式完整", "Agent 完整率"]
for col, h in enumerate(eval_headers, 1):
    c = ws_eval.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = eval_fill
    c.alignment = Alignment(horizontal="center")

eval_data = [
    ["单工具-数学计算", 15, 15, 15, 15, "100.0%", 15, 15, 15, "100.0%", 15, 15, 15, "100.0%"],
    ["单工具-其他工具", 15, 15, 15, 15, "100.0%", 15, 15, 15, "100.0%", 15, 14, 14, "93.3%"],
    ["多步链式调用",   20, 20, 20, 17, "85.0%",  20, 19, 13, "65.0%",  20, 20, 17, "85.0%"],
    ["工具选择能力",   15, 15, 15, 15, "100.0%", 14, 14, 14, "93.3%",  15, 15, 15, "100.0%"],
    ["中英文&表述变体", 15, 15, 14, 14, "93.3%",  15, 15, 15, "100.0%", 14, 14, 14, "93.3%"],
    ["合计",          80, 80, 79, 76, "95.0%",  79, 78, 72, "90.0%",  79, 78, 75, "93.8%"],
]
for row in eval_data:
    ws_eval.append(row)

for col in range(1, 15):
    ws_eval.column_dimensions[get_column_letter(col)].width = 16

# ==================== Sheet5: Think 评测结果 ====================
ws_think = wb.create_sheet("Think标签评测")
think_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
think_headers = ["问题", "权重", "Think输出", "思考质量", "答案正确"]
for col, h in enumerate(think_headers, 1):
    c = ws_think.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = think_fill
    c.alignment = Alignment(horizontal="center")

think_data = [
    ["小明有5个苹果，给了小红2个，又买了3个", "SFT", "有(冗长)", "反复纠结、逻辑混乱", "错误"],
    ["小明有5个苹果，给了小红2个，又买了3个", "GRPO", "有(垃圾)", "平衡平衡...重复文本(reward hacking)", "无意义输出"],
    ["小明有5个苹果，给了小红2个，又买了3个", "Agent", "有(简短)", "有推理但逻辑错", "错误 答案=3(正确=6)"],
    ["直角三角形两边3和4，第三条边？", "SFT", "有", "提到勾股定理但用错", "错误 答案=7(正确=5)"],
    ["直角三角形两边3和4，第三条边？", "GRPO", "有(垃圾)", "平衡平衡...重复文本(reward hacking)", "无意义输出"],
    ["直角三角形两边3和4，第三条边？", "Agent", "有(简短)", "有推理但结论错", "错误 答案=4(正确=5)"],
    ["水池进水3吨/时排水1吨/时，10吨水5小时后？", "SFT", "无", "未输出think标签", "错误 推理混乱(正确=20)"],
    ["水池进水3吨/时排水1吨/时，10吨水5小时后？", "GRPO", "有(垃圾)", "平衡平衡...重复文本(reward hacking)", "无意义输出"],
    ["水池进水3吨/时排水1吨/时，10吨水5小时后？", "Agent", "无", "未输出think标签", "错误 推理混乱(正确=20)"],
]
for row in think_data:
    ws_think.append(row)

ws_think.column_dimensions["A"].width = 40
ws_think.column_dimensions["B"].width = 10
ws_think.column_dimensions["C"].width = 15
ws_think.column_dimensions["D"].width = 35
ws_think.column_dimensions["E"].width = 28

# ==================== Sheet6: 训练曲线（matplotlib 生成图片嵌入） ====================
ws_chart = wb.create_sheet("训练曲线")

# 写入原始数据供参考
ws_chart.cell(row=1, column=1, value="PPO_Step").font = Font(bold=True)
ws_chart.cell(row=1, column=2, value="PPO_Reward").font = Font(bold=True)
ws_chart.cell(row=1, column=4, value="GRPO_Step").font = Font(bold=True)
ws_chart.cell(row=1, column=5, value="GRPO_Reward").font = Font(bold=True)

ppo_sampled = [r for r in ppo_data if r[0] % 10 == 0]
for i, r in enumerate(ppo_sampled):
    ws_chart.cell(row=i+2, column=1, value=r[0])
    ws_chart.cell(row=i+2, column=2, value=r[1])

for i, r in enumerate(grpo_data):
    ws_chart.cell(row=i+2, column=4, value=r[0])
    ws_chart.cell(row=i+2, column=5, value=r[1])

# --- 用 matplotlib 画 Reward 曲线 ---
def smooth(y, window=50):
    kernel = np.ones(window) / window
    return np.convolve(y, kernel, mode="valid")

ppo_steps = np.array([r[0] for r in ppo_sampled])
ppo_rewards = np.array([r[1] for r in ppo_sampled])
grpo_steps = np.array([r[0] for r in grpo_data])
grpo_rewards = np.array([r[1] for r in grpo_data])

fig, ax = plt.subplots(figsize=(14, 7))

# 原始数据（半透明）
ax.plot(ppo_steps, ppo_rewards, color="#4472C4", alpha=0.15, linewidth=0.5)
ax.plot(grpo_steps, grpo_rewards, color="#ED7D31", alpha=0.15, linewidth=0.5)

# 平滑曲线
w = 30
ppo_sm = smooth(ppo_rewards, w)
grpo_sm = smooth(grpo_rewards, w)
ax.plot(ppo_steps[w-1:], ppo_sm, color="#4472C4", linewidth=2.5, label="PPO Reward")
ax.plot(grpo_steps[w-1:], grpo_sm, color="#ED7D31", linewidth=2.5, label="GRPO Reward")

ax.set_xlabel("Training Step", fontsize=13)
ax.set_ylabel("Reward", fontsize=13)
ax.set_title("PPO vs GRPO - Reward Curve", fontsize=15, fontweight="bold")
ax.legend(fontsize=12, loc="upper left")
ax.set_xlim(0, 10000)
ax.grid(True, alpha=0.3)
fig.tight_layout()

chart_path = "eval_results/reward_curve.png"
fig.savefig(chart_path, dpi=150)
plt.close(fig)
print(f"Chart saved: {chart_path}")

img = XlImage(chart_path)
img.width = 900
img.height = 450
ws_chart.add_image(img, "G1")

out = "eval_results/training_summary.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: PPO({len(ppo_data)} steps), GRPO({len(grpo_data)} steps), Summary, ToolCall, Think, Chart")
